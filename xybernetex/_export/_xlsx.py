"""
XLSX renderer — converts an Artifact to an Excel workbook.
Requires: pip install xybernetex[xlsx]   (openpyxl)

Best results with artifacts whose content contains Markdown tables.
If no tables are found, the full content is written as plain text in
a single sheet.
"""
from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from xybernetex._models import Artifact

_MISSING = (
    "openpyxl is required for XLSX export.\n"
    "Install it with:  pip install xybernetex[xlsx]"
)


def render_xlsx(artifact: "Artifact", path: str) -> str:
    """
    Write *artifact* to an Excel workbook at *path*.

    Each Markdown table becomes a separate worksheet named after the
    nearest preceding heading.  Remaining content (paragraphs, lists,
    code) is collected in a 'Notes' sheet.

    If *path* is a directory the file is named after the artifact title.
    Returns the path of the written file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(_MISSING) from None

    from xybernetex._export._markdown import parse, inline_text
    from xybernetex._models import _safe_filename

    dest = _resolve_path(path, artifact.title, ".xlsx")
    dest.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    blocks = parse(artifact.content)

    # ── Colour palette ─────────────────────────────────────────────────────────
    HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
    ALT_FILL    = PatternFill("solid", fgColor="F0F0F8")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    BODY_FONT   = Font(name="Calibri", size=10)
    TITLE_FONT  = Font(name="Calibri", bold=True, size=14)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Collect tables and notes ───────────────────────────────────────────────
    tables: list[dict] = []   # {"name": str, "headers": [...], "rows": [...]}
    notes: list[str]   = []
    current_heading    = artifact.title

    for block in blocks:
        btype = block["type"]
        if btype == "heading":
            current_heading = block["text"]
        elif btype == "table":
            tables.append({
                "name": _sheet_name(current_heading, len(tables) + 1),
                "headers": block["headers"],
                "rows": block["rows"],
            })
        elif btype == "paragraph":
            notes.append(inline_text(block["text"]))
        elif btype in ("bullet_list", "ordered_list"):
            for item in block["items"]:
                notes.append(f"• {inline_text(item)}")
        elif btype == "code":
            notes.append(block["text"])
        elif btype == "hr":
            notes.append("─" * 40)

    # ── Write table sheets ─────────────────────────────────────────────────────
    for tbl in tables:
        ws = wb.create_sheet(title=tbl["name"])

        headers = tbl["headers"]
        rows    = tbl["rows"]
        col_count = max(len(headers), max((len(r) for r in rows), default=0))

        # Title row (row 1)
        ws.cell(1, 1, artifact.title).font = TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(col_count, 1))

        # Header row (row 2)
        for ci, h in enumerate(headers[:col_count], start=1):
            cell = ws.cell(2, ci, h)
            cell.font   = HEADER_FONT
            cell.fill   = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[2].height = 22

        # Data rows (start row 3)
        for ri, row_data in enumerate(rows, start=3):
            for ci, val in enumerate(row_data[:col_count], start=1):
                cell = ws.cell(ri, ci, val)
                cell.font   = BODY_FONT
                cell.fill   = ALT_FILL if ri % 2 == 0 else PatternFill()
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border

        # Auto column width
        for ci in range(1, col_count + 1):
            col_vals = [headers[ci - 1]] + [
                (r[ci - 1] if ci - 1 < len(r) else "") for r in rows
            ]
            max_len = min(max((len(str(v)) for v in col_vals), default=8), 50)
            ws.column_dimensions[get_column_letter(ci)].width = max_len + 2

        ws.freeze_panes = "A3"

    # ── Notes sheet ────────────────────────────────────────────────────────────
    if notes:
        ws = wb.create_sheet(title="Notes")
        ws.cell(1, 1, artifact.title).font = TITLE_FONT
        ws.column_dimensions["A"].width = 100
        for ri, line in enumerate(notes, start=2):
            cell = ws.cell(ri, 1, line)
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True)

    # Fallback: no tables AND no notes — write raw content
    if not tables and not notes:
        ws = wb.create_sheet(title="Content")
        ws.cell(1, 1, artifact.title).font = TITLE_FONT
        ws.column_dimensions["A"].width = 100
        for ri, line in enumerate(artifact.content.splitlines(), start=2):
            ws.cell(ri, 1, line).font = Font(name="Calibri", size=10)

    wb.save(str(dest))
    return str(dest)


# ── Helpers ────────────────────────────────────────────────────────────────────

def _resolve_path(path: str, title: str, ext: str) -> Path:
    from xybernetex._models import _safe_filename
    p = Path(path)
    if p.is_dir():
        p = p / f"{_safe_filename(title)}{ext}"
    return p


def _sheet_name(heading: str, index: int) -> str:
    """Truncate and sanitise a heading for use as an Excel sheet name."""
    illegal = r'\/:*?"<>|'
    name = "".join(c for c in heading if c not in illegal)[:28].strip() or f"Sheet{index}"
    return name
